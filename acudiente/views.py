from rest_framework import viewsets, status
from rest_framework.response import Response

#Documentacion
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
#Modelo
from .models import Acudiente
#Serializadores
from .serializers import AcudienteSerializer
#Autenticacion
from rest_framework.permissions import IsAuthenticated, AllowAny
#Permisos
from cuenta.permissions import IsEstudiante, IsProfesor, IsAdministrador, IsProfesorOrAdministrador, IsEstudianteOrAdministrador
#HTTP
from django.http import HttpResponse
#fecha y hora 
from datetime import datetime
#Acciones
from rest_framework.decorators import action
#Excel
import openpyxl
from openpyxl.styles import Font, Alignment

class AcudienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint para gestionar acudientes.
    
    Permite listar, crear, actualizar y eliminar acudientes.
    - Estudiantes: solo pueden crear acudientes
    - Profesores: no tienen acceso
    - Administradores: acceso completo
    """
    queryset = Acudiente.objects.all()
    serializer_class = AcudienteSerializer
    permission_classes = [IsAuthenticated]
    
    def get_permissions(self):
        """
        Define permisos según la acción solicitada:
        - create: Estudiantes y administradores pueden crear
        - list, retrieve, update, partial_update, destroy: Solo administradores
        """
        if self.action == 'create':
            #Cualquier usuario puedo crear un acudiente
            permission_classes = [AllowAny]
        elif self.action in ['update', 'partial_update']:
            # estudiantes pueden actualizar la informacion del acudiente
            permission_classes = [IsEstudianteOrAdministrador]
        else:
            # Solo administradores pueden listar, ver detalles, actualizar y eliminar
            permission_classes = [IsAdministrador]
        return [permission() for permission in permission_classes]

    @swagger_auto_schema(
        operation_summary="Listar todos los acudientes",
        operation_description="Retorna una lista de todos los acudientes registrados"
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    @swagger_auto_schema(
        operation_summary="Crear un acudiente",
        operation_description="Crea un nuevo registro de acudiente",
        responses={
            status.HTTP_201_CREATED: AcudienteSerializer,
            status.HTTP_400_BAD_REQUEST: "Datos de entrada inválidos"
        }
    )
    def create(self, request, *args, **kwargs):
        data = request.data

        # Verificar si ya existe un acudiente con el numero_documento_acudiente proporcionado
        numero_documento = data.get('numero_documento_acudiente')
        if numero_documento:
            existing_acudiente = Acudiente.objects.filter(numero_documento_acudiente=numero_documento).first()
            if existing_acudiente:
                serializer = self.get_serializer(existing_acudiente)
                return Response({"message": "Ya se encuentra registrado", "data": serializer.data}, status=status.HTTP_200_OK)

        #Crear el objeto usando el serializador
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        #Responder con los datos del nuevo acudiente
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    @swagger_auto_schema(
        operation_summary="Obtener un acudiente específico",
        operation_description="Retorna los detalles de un acudiente específico por su ID"
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)
    @swagger_auto_schema(
        operation_summary="Actualizar un acudiente",
        operation_description="Actualiza los datos de un acudiente existente",
        responses={
            status.HTTP_200_OK: AcudienteSerializer,
            status.HTTP_400_BAD_REQUEST: "Datos de entrada inválidos"
        }
    )
    def update(self, request, *args, **kwargs):
        data = request.data

        #Actualizar el objeto usando el serializador
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        #Responder con los datos del acudiente actualizado
        return Response(serializer.data, status=status.HTTP_200_OK)
    @swagger_auto_schema(
        operation_summary="Eliminar un estudiante",
        operation_description="Elimina permanentemente un estudiante del sistema"
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    
    @swagger_auto_schema(
        operation_summary="Exportar acudientes a Excel",
        operation_description="Genera un archivo Excel con la lista completa de acudientes",
        responses={
            status.HTTP_200_OK: "Archivo Excel generado correctamente",
            status.HTTP_500_INTERNAL_SERVER_ERROR: "Error al generar el archivo Excel"
        }
    )
    @action(detail=False, methods=['get'], url_path='export-excel',
            permission_classes=[IsAdministrador])
    def export_excel(self, request):
        # Obtener todos los acudientes
        acudientes = self.get_queryset()
        
        # Crear un libro de trabajo de Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Acudientes"
        
        # Definir encabezados basados en los campos del modelo
        headers = [
            'ID', 'Tipo Documento', 'Número Documento', 'Nombres', 
            'Apellidos', 'Teléfono', 'Correo Electrónico', 'Dirección'
        ]
        
        # Aplicar formato a encabezados
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
        # Agregar datos de acudientes
        for row_num, acudiente in enumerate(acudientes, 2):
            worksheet.cell(row=row_num, column=1).value = acudiente.id_acudiente
            worksheet.cell(row=row_num, column=2).value = acudiente.tipo_documento_acudiente
            worksheet.cell(row=row_num, column=3).value = acudiente.numero_documento_acudiente
            worksheet.cell(row=row_num, column=4).value = acudiente.nombre_acudiente
            worksheet.cell(row=row_num, column=5).value = acudiente.apellido_acudiente
            worksheet.cell(row=row_num, column=6).value = acudiente.celular_acudiente
            worksheet.cell(row=row_num, column=7).value = acudiente.email_acudiente
            
        # Ajustar ancho de columnas automáticamente
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        # Crear respuesta HTTP con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="acudientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        # Guardar el libro de trabajo en la respuesta
        workbook.save(response)
        return response


