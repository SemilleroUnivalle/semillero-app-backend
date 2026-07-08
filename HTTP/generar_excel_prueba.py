"""
Genera un archivo Excel de prueba para el endpoint de encuesta de satisfaccion.
Usa datos reales de la BD (estudiantes + modulos + docentes + monitores).
Se ejecuta dentro del contenedor Django.
"""
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'semillero_backend.settings')
django.setup()

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from inscripcion.models import Inscripcion

# ── 1. Obtener datos reales de la BD ──────────────────────────────────────────
print("Consultando inscripciones en la BD...")

inscripciones = (
    Inscripcion.objects
    .select_related(
        'id_estudiante',
        'id_modulo',
        'grupo__profesor',
    )
    .prefetch_related('id_modulo__monitor_academico')
    .order_by('id_estudiante__numero_documento')
    [:20]  # Tomamos hasta 20 para el archivo de prueba
)

filas = []
for insc in inscripciones:
    est     = insc.id_estudiante
    modulo  = insc.id_modulo
    grupo   = insc.grupo
    profesor = grupo.profesor if grupo else None
    monitor_obj = modulo.monitor_academico.first() if modulo else None

    nombre_docente = (
        f"{profesor.nombre} {profesor.apellido}" if profesor else "Sin docente"
    )
    nombre_monitor = (
        f"{monitor_obj.nombre} {monitor_obj.apellido}" if monitor_obj else "Sin monitor"
    )
    nombre_modulo = modulo.nombre_modulo if modulo else "Sin módulo"

    filas.append({
        "documento":       est.numero_documento,
        "nombre":          f"{est.nombre} {est.apellido}",
        "modulo":          nombre_modulo,
        "docente":         nombre_docente,
        "monitor":         nombre_monitor,
        "nota_modulo":     None,
        "nota_docente":    None,
        "nota_monitor":    None,
        "nota_estudiante": None,
    })

print(f"  → {len(filas)} inscripciones encontradas.")

# ── 2. Crear el workbook ──────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Encuesta de Satisfaccion"

COLUMNAS = [
    ("documento",       "Documento",            16),
    ("nombre",          "Nombre completo",       28),
    ("modulo",          "Módulo",                22),
    ("docente",         "Docente",               22),
    ("monitor",         "Monitor",               22),
    ("nota_modulo",     "Nota módulo",           13),
    ("nota_docente",    "Nota docente",          13),
    ("nota_monitor",    "Nota monitor",          13),
    ("nota_estudiante", "Autoevaluación",        14),
]

# ── Estilos ───────────────────────────────────────────────────────────────────
AZUL_OSCURO   = "1F3864"
AZUL_CLARO    = "BDD7EE"
AMARILLO      = "FFD966"
BLANCO        = "FFFFFF"
GRIS_CLARO    = "F2F2F2"
VERDE_CLARO   = "E2EFDA"

header_font    = Font(name="Calibri", bold=True, color=BLANCO, size=11)
subheader_font = Font(name="Calibri", bold=True, color="1F3864", size=10)
data_font      = Font(name="Calibri", size=10)
nota_font      = Font(name="Calibri", size=10, color="1F3864")

header_fill    = PatternFill("solid", fgColor=AZUL_OSCURO)
nota_fill      = PatternFill("solid", fgColor=AMARILLO)
datos_fill1    = PatternFill("solid", fgColor=BLANCO)
datos_fill2    = PatternFill("solid", fgColor=GRIS_CLARO)

center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
thick= Side(style="medium", color=AZUL_OSCURO)
border_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header= Border(left=thick, right=thick, top=thick, bottom=thick)

# ── Fila 1: Título ─────────────────────────────────────────────────────────────
ws.merge_cells(f"A1:{get_column_letter(len(COLUMNAS))}1")
titulo_cell = ws["A1"]
titulo_cell.value     = "ENCUESTA DE SATISFACCIÓN – SEMILLERO UNIVALLE"
titulo_cell.font      = Font(name="Calibri", bold=True, size=14, color=BLANCO)
titulo_cell.fill      = header_fill
titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Fila 2: Instrucción ────────────────────────────────────────────────────────
ws.merge_cells(f"A2:{get_column_letter(len(COLUMNAS))}2")
instr_cell = ws["A2"]
instr_cell.value = (
    "Instrucciones: Rellena las columnas de color amarillo con notas de 0.0 a 5.0  |  "
    "Las columnas en gris (documento, nombre, módulo, docente, monitor) son informativas y NO se modifican."
)
instr_cell.font      = Font(name="Calibri", italic=True, size=9, color="595959")
instr_cell.fill      = PatternFill("solid", fgColor="DEEAF1")
instr_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 20

# ── Fila 3: Encabezados ────────────────────────────────────────────────────────
for col_idx, (campo, etiqueta, ancho) in enumerate(COLUMNAS, start=1):
    cell = ws.cell(row=3, column=col_idx, value=etiqueta)
    # Columnas de nota → fondo amarillo
    if campo.startswith("nota_"):
        cell.fill = nota_fill
        cell.font = Font(name="Calibri", bold=True, color="1F3864", size=11)
    else:
        cell.fill = header_fill
        cell.font = header_font
    cell.alignment = center_align
    cell.border    = border_thin
    ws.column_dimensions[get_column_letter(col_idx)].width = ancho

ws.row_dimensions[3].height = 36

# ── Filas 4+: Datos ───────────────────────────────────────────────────────────
CAMPOS_ORDEN = [c[0] for c in COLUMNAS]
CAMPOS_INFO  = {"documento", "nombre", "modulo", "docente", "monitor"}
CAMPOS_NOTA  = {"nota_modulo", "nota_docente", "nota_monitor", "nota_estudiante"}

for row_idx, fila in enumerate(filas, start=4):
    fill = datos_fill1 if (row_idx % 2 == 0) else datos_fill2

    for col_idx, campo in enumerate(CAMPOS_ORDEN, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=fila[campo])
        cell.border    = border_thin

        if campo in CAMPOS_INFO:
            cell.font      = data_font
            cell.fill      = fill
            cell.alignment = left_align if campo != "documento" else center_align
        else:
            # Celda de nota → amarilla clara con hint de formato
            cell.fill      = PatternFill("solid", fgColor="FFFACD")
            cell.font      = Font(name="Calibri", size=10, color="BF8F00")
            cell.alignment = center_align
            cell.number_format = "0.0"

    ws.row_dimensions[row_idx].height = 18

# ── Agregar fila de ejemplo si no hay datos reales ────────────────────────────
if not filas:
    print("  ⚠  No hay inscripciones en BD. Agregando fila de ejemplo.")
    row_ejemplo = 4
    ejemplos = ["1001234567", "Juan Pérez", "Matemáticas", "Carlos Gómez", "Laura Díaz", 4.5, 4.7, 4.3, 4.8]
    for col_idx, valor in enumerate(ejemplos, start=1):
        cell = ws.cell(row=row_ejemplo, column=col_idx, value=valor)
        cell.border    = border_thin
        cell.alignment = center_align
        cell.font      = data_font
    ws.row_dimensions[row_ejemplo].height = 18

# ── Anclar filas de encabezado ────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Segunda hoja: instrucciones detalladas ─────────────────────────────────────
ws2 = wb.create_sheet("Instrucciones")

instrucciones = [
    ("COLUMNA",          "DESCRIPCIÓN",                                    "EDITABLE?"),
    ("documento",        "Número de documento del estudiante  ← KEY",      "NO"),
    ("nombre",           "Nombre y apellido (referencia visual)",           "NO"),
    ("modulo",           "Nombre del módulo cursado",                       "NO"),
    ("docente",          "Profesor asignado al grupo",                      "NO"),
    ("monitor",          "Monitor académico del módulo",                    "NO"),
    ("nota_modulo",      "Calificación al contenido del módulo  (0.0–5.0)","SÍ"),
    ("nota_docente",     "Calificación al desempeño del docente (0.0–5.0)","SÍ"),
    ("nota_monitor",     "Calificación al desempeño del monitor (0.0–5.0)","SÍ"),
    ("nota_estudiante",  "Autoevaluación del estudiante         (0.0–5.0)","SÍ"),
    ("",                 "",                                                ""),
    ("ENDPOINT",         f"POST /encuesta_satisfaccion/encuesta/cargar-excel/",  ""),
    ("CAMPO FORM",       "archivo  (multipart/form-data)",                  ""),
    ("FORMATO",          ".xlsx  o  .xls",                                  ""),
    ("ESCALAS",          "Notas de 0.0 a 5.0  (decimales con punto o coma)",""),
    ("RESPUESTA OK",     "HTTP 200  →  { total_filas, creadas, actualizadas, errores, detalle_exito, detalle_errores }", ""),
]

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 12

for r, row_data in enumerate(instrucciones, start=1):
    for c, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = Font(bold=True, color=BLANCO)
            cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = border_thin

# ── Guardar ────────────────────────────────────────────────────────────────────
output_path = "/tmp/encuesta_satisfaccion_prueba.xlsx"
wb.save(output_path)
print(f"\n✅ Archivo generado: {output_path}")
print(f"   Filas de datos   : {len(filas)}")
print(f"   Columnas         : {len(COLUMNAS)}")
